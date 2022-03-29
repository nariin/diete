# from django.apps import AppConfig
import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "server.settings")
import django
django.setup()
from django.core.wsgi import get_wsgi_application
import threading
import time
from openpyxl import Workbook, load_workbook
from google_images_download import google_images_download  
from menus.serializers import FoodSerializer

FILE_NAME = "datas/통합 식품영양성분DB_음식_20220308.xlsx"
SHEET_NAME = "Data"

# 엑셀 파일의 필요 속성 13개
xls_properties = ["식품명"]

# Food 테이블의 13개 속성 중 image를 제외한 12가지 매칭 + 품목대표인지 확인을 위해 "상용제품" 속성 추가 필요 -> "상용제품" : commercialFood
table_properties = {"식품명" : "foodName"}

class XlsxParser():
    def __init__(self):
        self.load_wb = load_workbook(filename= FILE_NAME, data_only=True)
        self.load_ws = self.load_wb[SHEET_NAME]
        self.match_Properties_To_Index()
        print("data amount :", (self.load_ws.max_row-4))

    def match_Properties_To_Index(self):
        # 엑셀 속성의 컬럼 index와 속성명을 매칭
        self.xls_index_to_values = {}
        for tuples in self.load_ws['4']:
            if tuples.value in xls_properties:
                self.xls_index_to_values[tuples.column] = tuples.value
    
    def input_datas(self):
        cnt = 0;
        fail_cnt = 0;
        food_set = set()
        foodString = ""
        for row in self.load_ws.rows:
            if row[0].row < 5: continue
            for cell in row:
                if cell.column == 5 and cell.value == "외식":
                    fail_cnt += 1
                    break
                if cell.column == 6:
                    if cell.value.strip() in food_set: break
                    cnt += 1
                    print(cell.value.strip())
                    foodString += cell.value.strip() + ","
                    if(cnt%50 == 0):
                        try:
                            threading.Thread(target=self.crawlImages, args=(foodString[:-1],)).start()
                            # self.crawlImages(cell.value.strip())
                            time.sleep(0.05)
                        except:
                            print("에러!! :", cell.value.strip())
                            pass
                        foodString = ""
                    print("현재 갯수 :", cnt)
                    food_set.add(cell.value.strip())

        return cnt, fail_cnt
    
    def crawlImages(self, foodString):
        response = google_images_download.googleimagesdownload()  
        arguments = {"keywords": foodString
        , "limit":1, "print_urls":False, "format":"jpg", "size":"medium"
        , "no_directory":True, "no_numbering":True}

        path = response.download(arguments)[0]

        for curr in path.keys():
            old_path = path[curr][0]
            old_slash = old_path.rfind("\\")
            old_file = old_path[old_slash+1:]
            old_path = old_path[:old_slash]
            file_oldname = os.path.join(old_path, old_file)
            # print("경로 :", old_path)
            next_file = curr + ".jpg"
            file_newname = os.path.join(old_path, next_file)

            os.rename(file_oldname, file_newname)

def main():
    print("start image_crawler")

    print("initializing image_crawler")
    xlsxParser = XlsxParser()
    print("finished")

    xlsxParser.match_Properties_To_Index()

    print("start input_datas")
    cnt, fail_cnt = xlsxParser.input_datas()
    print("finished")

    print("-----------------result-----------------")
    print("품목대표 :", (cnt), ", 외식 :", fail_cnt, ", Total :", (cnt+fail_cnt))
    # print("Successed :", (cnt), ", Failed :", fail_cnt, ", Total :", (cnt+fail_cnt))
    print("----------------------------------------")

    print("finished image_crawler")


if __name__ == '__main__':
    main()
